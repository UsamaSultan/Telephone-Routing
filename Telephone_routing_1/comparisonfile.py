""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
""" Reading from file and Reccomending cheapest operator for specific number """

import xlrd
import sys
import difflib
import openpyxl
import tkinter
from tkinter import *
from tkinter import filedialog as fd

main_win = tkinter.Tk()

file_name = fd.askopenfilename(parent=main_win, initialdir="/", title='Please select a file')

main_win.withdraw()

class ala_comparsion():    

    def comparison(inp):
        
        ##### Checking if the input is a number or not #####
        
        try:
            
            inpcheck = int(inp)
            
        except:
                    
            return (print("Not a valid number, Please dial the number again"))
        
            sys.exit()
        
            
        Operator = []
        
        Code = []
        
        PPU = []
        
        
        #file_name = "D:\Programming_Exercise\input.xlsx"
        wb = xlrd.open_workbook(file_name)

        sheet = wb.sheet_by_index(0)

        
        
        for i in range(1, sheet.nrows):
            
            Operator.append(sheet.cell(i, 0).value)
            
            Code.append(int(sheet.cell(i, 1).value))
            
            PPU.append(sheet.cell(i, 2).value)
            
            
        Numbersplit = []
        
        value = ""
        
        for i in range(len(inp)):
            
            value = value + inp[i]
            
            Numbersplit.append(value)
            
        
        
        ##### COMPARISON #####
        
        el=0
        
        CodeComparison = []
        
        pricecheck = []
        
        for i in range(len(Numbersplit)):
            
            for j in range(len(Code)):
                
                ratioScore = difflib.SequenceMatcher(None, str(Numbersplit[i]), str(Code[j])).ratio()

                if ratioScore == 1.0 and len(str(Code[j])) == el:
                    
                    el = len(str(Code[j]))
                    
                    CodeComparison.append(j)
                    
                    pricecheck.append(PPU[j])

                if ratioScore == 1.0 and len(str(Code[j])) > el:
                    
                    el = len(str(Code[j]))
                    
                    CodeComparison.clear()
                    
                    pricecheck.clear()
                    
                    CodeComparison.append(j)
                    
                    pricecheck.append(PPU[j])
                    
         
         
                    
        ##### CALCULATING RESULTS #####
        
        try:
            
            minprice = pricecheck.index(min(pricecheck))
            
            OperatorResult = Operator[CodeComparison[minprice]]
            
            coderesult = Code[CodeComparison[minprice]]
            
            PPUresult = PPU[CodeComparison[minprice]]
            
            return(print("\n",OperatorResult,"\n Price : $"+str(PPUresult)+"/min","\n Code : ",coderesult,"\n"))
        
        except:
            
            return(print("Dialing number is invalid i.e. Dialing Code is not in any of our operator's input"))
            
            sys.exit()
            
            
        ##### CLEARING THE CACHES #####
        
        Numbersplit.clear()
        
        CodeComparison.clear()
        
        pricecheck.clear()
